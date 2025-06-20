from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from bloodpoint_app.models import campana, TIPO_SANGRE_CHOICES

def generar_excel_campana(campana_id, response):
    campana_obj = campana.objects.get(id_campana=campana_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Campaña"

    # Estilos
    title_font = Font(size=16, bold=True, color="FF0000")
    subtitle_font = Font(size=12, italic=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FF0000")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Título principal
    nombre_representante = campana_obj.id_representante.full_name()
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Campaña: {campana_obj.nombre_campana} - Representante: {nombre_representante}"
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    # Subtítulo con fechas
    fecha_inicio = campana_obj.fecha_campana.strftime('%Y-%m-%d')
    fecha_termino = campana_obj.fecha_termino.strftime('%Y-%m-%d') if campana_obj.fecha_termino else "N/A"
    ws.merge_cells("A2:J2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"Desde {fecha_inicio} hasta {fecha_termino}"
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = center_alignment

    # Encabezados de la tabla de resumen
    headers = [
        "ID Campaña",
        "Nombre Campaña",
        "Fecha Inicio",
        "Fecha Término",
        "Meta Donaciones (unidades)",
        "Donaciones Realizadas (unidades)",
        "Porcentaje Cumplimiento (%)",
        "Total ML Donados",
        "Estado Campaña",
        "Representante Responsable",
    ]
    ws.append([])  # Fila vacía para separar visualmente
    ws.append(headers)

    header_row_idx = 4
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Datos resumen
    total_donaciones = campana_obj.donacion_set.count()
    total_ml = sum(d.cantidad_donacion for d in campana_obj.donacion_set.all())
    meta = int(campana_obj.meta) if campana_obj.meta else 0
    porcentaje = (total_donaciones / meta) * 100 if meta else 0

    ws.append([
        campana_obj.id_campana,
        campana_obj.nombre_campana,
        fecha_inicio,
        fecha_termino,
        meta,
        total_donaciones,
        f"{porcentaje:.1f}%",
        total_ml,
        campana_obj.estado,
        nombre_representante,
    ])

    # Estilos para la fila de datos
    data_row_idx = header_row_idx + 1
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=data_row_idx, column=col_num)
        cell.alignment = center_alignment
        cell.border = thin_border

    # Segunda hoja: ML por tipo de sangre
    ws2 = wb.create_sheet(title="ML por Tipo de Sangre")
    sub_headers = ["Tipo de Sangre", "Total ML Donados"]
    ws2.append(sub_headers)
    for col_num, header in enumerate(sub_headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for row_num, (tipo, _) in enumerate(TIPO_SANGRE_CHOICES, 2):
        total_ml_tipo = sum(
            d.cantidad_donacion
            for d in campana_obj.donacion_set.filter(id_donante__tipo_sangre=tipo)
        )
        ws2.append([tipo, total_ml_tipo])
        for col_num in range(1, 3):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = thin_border

    # Ajustar anchos de columna (evitar MergedCell error)
    for sheet in [ws, ws2]:
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 4

    # Guardar archivo en response
    wb.save(response)
    if hasattr(response, "seek"):
        response.seek(0)
