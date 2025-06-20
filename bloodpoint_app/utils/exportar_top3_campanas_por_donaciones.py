from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from bloodpoint_app.models import campana, donacion

def exportar_top3_campanas_por_donaciones():
    campañas_data = []

    for c in campana.objects.all():
        donaciones = donacion.objects.filter(campana_relacionada=c)
        total_donaciones = donaciones.count()
        total_validadas = donaciones.filter(validada=True).count()
        total_intencionadas = donaciones.filter(es_intencion=True).count()
        total_sangre = sum([d.cantidad_donacion for d in donaciones])
        nuevos_donantes = sum([1 for d in donaciones if d.id_donante.nuevo_donante])

        porcentaje_nuevos = (nuevos_donantes / total_donaciones * 100) if total_donaciones else 0

        try:
            meta_personas = int(c.meta) if c.meta else 0
        except (ValueError, TypeError):
            meta_personas = 0

        porcentaje_meta = (total_donaciones / meta_personas * 100) if meta_personas else 0

        campañas_data.append({
            "campana": c,
            "total_donaciones": total_donaciones,
            "validadas": total_validadas,
            "intencionadas": total_intencionadas,
            "total_sangre": total_sangre,
            "meta": meta_personas,
            "porc_meta": porcentaje_meta,
            "porc_nuevos": porcentaje_nuevos
        })

    top3 = sorted(campañas_data, key=lambda x: x["total_donaciones"], reverse=True)[:3]

    wb = Workbook()
    ws = wb.active
    ws.title = "Top 3 campañas"

    # Estilos
    burdeo = "8B0000"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=burdeo)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Título
    titulo = "Reporte TOP 3 Campañas de donación a nivel nacional"
    fecha = datetime.now().strftime("Generado el %d-%m-%Y a las %H:%M")

    ws.merge_cells("A1:N1")
    cell_titulo = ws["A1"]
    cell_titulo.value = titulo
    cell_titulo.font = Font(size=16, bold=True, color="000000")  # negro
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:N2")
    cell_fecha = ws["A2"]
    cell_fecha.value = fecha
    cell_fecha.font = Font(italic=True, color="555555")
    cell_fecha.alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = [
        "ID campaña", "Nombre", "Fecha inicio", "Fecha fin", "Comuna", "Centro",
        "Representante", "Meta (personas)", "Donaciones recibidas", "% meta alcanzada",
        "Validas", "Intencionadas", "Total sangre (ml)", "% nuevos donantes"
    ]
    ws.append([])  # fila 3 vacía
    ws.append(headers)  # fila 4

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Datos desde fila 5
    for i, data in enumerate(top3, start=5):
        c = data["campana"]
        fila = [
            c.id_campana,
            c.nombre_campana,
            c.fecha_campana,
            c.fecha_termino,
            c.id_centro.comuna if c.id_centro else "",
            c.id_centro.nombre_centro if c.id_centro else "",
            f"{c.id_representante.nombre} {c.id_representante.apellido}" if c.id_representante else "",
            data["meta"],
            data["total_donaciones"],
            round(data["porc_meta"], 2),
            data["validadas"],
            data["intencionadas"],
            data["total_sangre"],
            round(data["porc_nuevos"], 2)
        ]
        for j, valor in enumerate(fila, 1):
            cell = ws.cell(row=i, column=j, value=valor)
            cell.border = border
            cell.alignment = center_align

    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = length + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
